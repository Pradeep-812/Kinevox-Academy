"""
utils/email_notify.py
======================
Email notification helper for Kinevox Academy.

Sends an email with the new record data as an Excel (.xlsx) attachment
whenever a new Employee or Student is created.

Configuration (via .env):
  NOTIFY_EMAIL_TO       — recipient address (required)
  SMTP_HOST             — SMTP server host  (default: smtp.gmail.com)
  SMTP_PORT             — SMTP port         (default: 587)
  SMTP_USER             — SMTP login user   (required)
  SMTP_PASSWORD         — SMTP password / app-password (required)
  NOTIFY_FROM_NAME      — Friendly sender name (default: Kinevox Academy)

Usage:
  from utils.email_notify import notify_new_record
  notify_new_record("employee", emp_dict)   # fire-and-forget (threaded)
  notify_new_record("student",  stu_dict)
"""

import io
import mimetypes
import os
import smtplib
import threading
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


# ── Configuration helpers ─────────────────────────────────────────────────────

def _cfg(key: str, default: str = "") -> str:
    return os.getenv(key, default).strip()


# ── Excel builder ─────────────────────────────────────────────────────────────

def _build_excel(record_type: str, data: dict) -> bytes:
    """Return an .xlsx file as bytes containing the record fields."""
    if not _HAS_OPENPYXL:
        raise RuntimeError("openpyxl is not installed — cannot build Excel attachment")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = record_type.capitalize()

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="6C63FF")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    label_font   = Font(bold=True, color="333333", size=10)
    value_font   = Font(color="111111", size=10)
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ─────────────────────────────────────────────────────────────
    title_text = (
        f"New {'Employee' if record_type == 'employee' else 'Student'} Record"
        f" — Kinevox Academy"
    )
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font  = Font(bold=True, color="FFFFFF", size=13)
    title_cell.fill  = header_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Timestamp row ─────────────────────────────────────────────────────────
    ws.merge_cells("A2:B2")
    ts_cell = ws["A2"]
    ts_cell.value = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ts_cell.font  = Font(italic=True, color="888888", size=9)
    ts_cell.alignment = center
    ws.row_dimensions[2].height = 18

    # Spacer
    ws.row_dimensions[3].height = 8

    # ── Column headers ────────────────────────────────────────────────────────
    ws["A4"].value = "Field"
    ws["B4"].value = "Value"
    for col in ("A4", "B4"):
        cell = ws[col]
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border
    ws.row_dimensions[4].height = 22

    # ── Field rows ────────────────────────────────────────────────────────────
    SKIP = {"photo", "sign", "idDoc", "resume", "courses"}

    # Human-friendly field name mapping
    LABELS = {
        # Employee
        "id":          "Employee ID",
        "name":        "Full Name",
        "role":        "Role / Designation",
        "dept":        "Department",
        "email":       "Email Address",
        "phone":       "Phone Number",
        "doj":         "Date of Joining",
        "salary":      "Basic Salary (₹)",
        "bonus":       "Bonus (₹)",
        "credits":     "Credits",
        "schedule":    "Schedule",
        "status":      "Status",
        # Student
        "institution": "Institution",
        "course":      "Course",
        "assigned_to": "Trainer ID",
        "trainer":     "Trainer Name",
        "trainer_name":"Trainer Name",
        "enroll_date": "Enroll Date",
        "progress":    "Progress (%)",
        "percentage":  "Completion (%)",
        "dob":         "Date of Birth",
    }

    row = 5
    for key, value in data.items():
        if key in SKIP:
            continue
        if value is None or value == "":
            continue

        label = LABELS.get(key, key.replace("_", " ").title())
        val   = str(value)

        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font      = label_font
        label_cell.alignment = left
        label_cell.border    = border

        value_cell = ws.cell(row=row, column=2, value=val)
        value_cell.font      = value_font
        value_cell.alignment = left
        value_cell.border    = border

        ws.row_dimensions[row].height = 20
        row += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Email sender ──────────────────────────────────────────────────────────────

def _send_email(record_type: str, data: dict) -> None:
    """Build and send the notification email (blocking — run in a thread)."""
    to_addr   = _cfg("NOTIFY_EMAIL_TO")
    smtp_host = _cfg("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(_cfg("SMTP_PORT", "587"))
    smtp_user = _cfg("SMTP_USER")
    smtp_pass = _cfg("SMTP_PASSWORD")
    from_name = _cfg("NOTIFY_FROM_NAME", "Kinevox Academy")

    if not all([to_addr, smtp_user, smtp_pass]):
        print("[email_notify] NOTIFY_EMAIL_TO / SMTP_USER / SMTP_PASSWORD not set — skipping.")
        return

    # ── Build email ───────────────────────────────────────────────────────────
    label      = "Employee" if record_type == "employee" else "Student"
    rec_name   = data.get("name", "Unknown")
    rec_id     = data.get("id", "—")
    subject    = f"[Kinevox Academy] New {label} Added: {rec_name} ({rec_id})"
    timestamp  = datetime.now().strftime("%d %b %Y at %I:%M %p")

    # Plain-text fallback
    plain_lines = [f"New {label} record added on {timestamp}.", ""]
    for k, v in data.items():
        if k in {"photo", "sign", "idDoc", "resume", "courses"} or v is None or v == "":
            continue
        plain_lines.append(f"  {k}: {v}")
    plain_lines += ["", "—", "Kinevox Academy Management System"]
    plain_body = "\n".join(plain_lines)

    # HTML body
    rows_html = ""
    SKIP = {"photo", "sign", "idDoc", "resume", "courses"}
    LABELS = {
        "id":"ID","name":"Full Name","role":"Role","dept":"Department",
        "email":"Email","phone":"Phone","doj":"Date of Joining",
        "salary":"Basic Salary (₹)","bonus":"Bonus (₹)","credits":"Credits",
        "schedule":"Schedule","status":"Status","institution":"Institution",
        "course":"Course","assigned_to":"Trainer ID","trainer":"Trainer",
        "trainer_name":"Trainer Name","enroll_date":"Enroll Date",
        "progress":"Progress (%)","percentage":"Completion (%)","dob":"Date of Birth",
    }
    for k, v in data.items():
        if k in SKIP or v is None or v == "":
            continue
        friendly = LABELS.get(k, k.replace("_", " ").title())
        rows_html += (
            f'<tr><td style="padding:9px 14px;font-weight:600;color:#444;'
            f'background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">'
            f'{friendly}</td>'
            f'<td style="padding:9px 14px;color:#222;border-bottom:1px solid #eee">'
            f'{v}</td></tr>'
        )

    html_body = f"""
<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f0f0f7;font-family:'DM Sans',Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td align="center" style="padding:32px 16px">
<table width="560" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:12px;overflow:hidden;
         box-shadow:0 4px 24px rgba(0,0,0,.08)">

  <!-- Header -->
  <tr>
    <td style="background:linear-gradient(135deg,#6c63ff,#38bdf8);
               padding:28px 32px;text-align:center">
      <div style="font-size:22px;font-weight:800;color:#fff;letter-spacing:-.5px">
        🎓 Kinevox Academy
      </div>
      <div style="color:rgba(255,255,255,.85);font-size:13px;margin-top:4px">
        Management System Notification
      </div>
    </td>
  </tr>

  <!-- Alert banner -->
  <tr>
    <td style="background:#f0edff;padding:14px 32px;border-bottom:2px solid #6c63ff">
      <span style="font-size:15px;font-weight:700;color:#6c63ff">
        ✅ New {label} Record Added
      </span>
      <span style="font-size:12px;color:#888;float:right;margin-top:2px">
        {timestamp}
      </span>
    </td>
  </tr>

  <!-- Data table -->
  <tr>
    <td style="padding:24px 32px">
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border:1px solid #e8e8f0;border-radius:8px;overflow:hidden">
        {rows_html}
      </table>
    </td>
  </tr>

  <!-- Footer -->
  <tr>
    <td style="background:#f8f8fc;padding:16px 32px;text-align:center;
               font-size:11px;color:#aaa;border-top:1px solid #eee">
      The full record is attached as an Excel file for your records.<br>
      This is an automated message from Kinevox Academy Management System.
    </td>
  </tr>

</table>
</td></tr>
</table>
</body>
</html>
"""

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = f"{from_name} <{smtp_user}>"
    msg["To"]      = to_addr

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(plain_body, "plain"))
    alt.attach(MIMEText(html_body,  "html"))
    msg.attach(alt)

    # ── Attach Excel ──────────────────────────────────────────────────────────
    try:
        xlsx_bytes = _build_excel(record_type, data)
        filename   = f"kinevox_{record_type}_{rec_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        part = MIMEBase("application",
                        "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(xlsx_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(part)
    except Exception as exc:
        print(f"[email_notify] Excel build failed: {exc} — sending without attachment.")

    # ── Send ──────────────────────────────────────────────────────────────────
    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, [to_addr], msg.as_string())
        print(f"[email_notify] ✅ Notification sent to {to_addr} for {record_type} {rec_id}")
    except Exception as exc:
        print(f"[email_notify] ❌ Failed to send email: {exc}")


# ── Public API ────────────────────────────────────────────────────────────────

def notify_new_record(record_type: str, data: dict) -> None:
    """
    Fire-and-forget: sends an email notification in a background thread.

    Args:
        record_type: "employee" or "student"
        data:        The record dict (same as what was saved to DB)
    """
    t = threading.Thread(
        target=_send_email,
        args=(record_type, data),
        daemon=True,
        name=f"email-notify-{record_type}",
    )
    t.start()


# ── Document upload notification ───────────────────────────────────────────────

def _send_document_email(
    record_type: str,
    record: dict,
    doc_type_label: str,
    filepath: str,
    original_filename: str,
) -> None:
    """Send an email with the uploaded document as an attachment."""
    to_addr   = _cfg("NOTIFY_EMAIL_TO")
    smtp_host = _cfg("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(_cfg("SMTP_PORT", "587"))
    smtp_user = _cfg("SMTP_USER")
    smtp_pass = _cfg("SMTP_PASSWORD")
    from_name = _cfg("NOTIFY_FROM_NAME", "Kinevox Academy")

    if not all([to_addr, smtp_user, smtp_pass]):
        print("[email_notify] NOTIFY_EMAIL_TO / SMTP_USER / SMTP_PASSWORD not set — skipping.")
        return

    label    = "Employee" if record_type == "employee" else "Student"
    rec_name = record.get("name", "Unknown")
    rec_id   = record.get("id", "—")
    dept     = record.get("dept") or record.get("course") or "—"
    role     = record.get("role") or record.get("institution") or "—"
    timestamp = datetime.now().strftime("%d %b %Y at %I:%M %p")

    subject = (
        f"[Kinevox Academy] Document Upload: {doc_type_label} "
        f"— {rec_name} ({rec_id})"
    )

    plain_body = (
        f"A new document was uploaded for {label}: {rec_name} ({rec_id})\n"
        f"Document Type : {doc_type_label}\n"
        f"File          : {original_filename}\n"
        f"Uploaded at   : {timestamp}\n\n"
        f"The document is attached to this email.\n"
        f"— Kinevox Academy Management System"
    )

    html_body = f"""<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f0f0f7;font-family:'DM Sans',Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td align="center" style="padding:32px 16px">
<table width="560" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:12px;overflow:hidden;
         box-shadow:0 4px 24px rgba(0,0,0,.08)">
  <tr>
    <td style="background:linear-gradient(135deg,#6c63ff,#38bdf8);padding:28px 32px;text-align:center">
      <div style="font-size:22px;font-weight:800;color:#fff">🎓 Kinevox Academy</div>
      <div style="color:rgba(255,255,255,.85);font-size:13px;margin-top:4px">Document Upload Notification</div>
    </td>
  </tr>
  <tr>
    <td style="background:#f0edff;padding:14px 32px;border-bottom:2px solid #6c63ff">
      <span style="font-size:15px;font-weight:700;color:#6c63ff">📎 New Document Uploaded</span>
      <span style="font-size:12px;color:#888;float:right;margin-top:2px">{timestamp}</span>
    </td>
  </tr>
  <tr>
    <td style="padding:24px 32px">
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border:1px solid #e8e8f0;border-radius:8px;overflow:hidden">
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">{label} Name</td>
          <td style="padding:10px 14px;color:#222;border-bottom:1px solid #eee">{rec_name}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">{label} ID</td>
          <td style="padding:10px 14px;color:#222;border-bottom:1px solid #eee">{rec_id}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">{'Department' if record_type=='employee' else 'Institution'}</td>
          <td style="padding:10px 14px;color:#222;border-bottom:1px solid #eee">{dept}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;border-bottom:1px solid #eee;white-space:nowrap">Document Type</td>
          <td style="padding:10px 14px;color:#222;border-bottom:1px solid #eee">{doc_type_label}</td>
        </tr>
        <tr>
          <td style="padding:10px 14px;font-weight:600;color:#444;background:#f8f8fc;white-space:nowrap">File Name</td>
          <td style="padding:10px 14px;color:#222">{original_filename}</td>
        </tr>
      </table>
      <div style="margin-top:16px;padding:14px 18px;background:#f0fff4;border:1px solid #6ee7b7;
                  border-radius:8px;font-size:13px;color:#065f46">
        📎 The uploaded document is attached to this email.
      </div>
    </td>
  </tr>
  <tr>
    <td style="background:#f8f8fc;padding:16px 32px;text-align:center;
               font-size:11px;color:#aaa;border-top:1px solid #eee">
      This is an automated message from Kinevox Academy Management System.
    </td>
  </tr>
</table>
</td></tr>
</table>
</body>
</html>"""

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = f"{from_name} <{smtp_user}>"
    msg["To"]      = to_addr

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(plain_body, "plain"))
    alt.attach(MIMEText(html_body,  "html"))
    msg.attach(alt)

    # Attach the actual uploaded file
    try:
        mime_type, _ = mimetypes.guess_type(filepath)
        main_type, sub_type = (mime_type or "application/octet-stream").split("/", 1)
        with open(filepath, "rb") as fh:
            part = MIMEBase(main_type, sub_type)
            part.set_payload(fh.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment",
                        filename=original_filename)
        msg.attach(part)
    except Exception as exc:
        print(f"[email_notify] Could not attach file: {exc}")

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, [to_addr], msg.as_string())
        print(f"[email_notify] ✅ Document email sent for {record_type} {rec_id} — {doc_type_label}")
    except Exception as exc:
        print(f"[email_notify] ❌ Failed to send document email: {exc}")


def notify_document_upload(
    record_type: str,
    record: dict,
    doc_type_label: str,
    filepath: str,
    original_filename: str,
) -> None:
    """Fire-and-forget document upload email notification."""
    t = threading.Thread(
        target=_send_document_email,
        args=(record_type, record, doc_type_label, filepath, original_filename),
        daemon=True,
        name="email-notify-doc",
    )
    t.start()


# ── Batch document upload notification ────────────────────────────────────────

def _send_documents_batch_email(
    record_type: str,
    record: dict,
    documents: list,   # list of (doc_type_label, filepath, original_filename)
) -> None:
    """Send one email with ALL uploaded documents as attachments."""
    to_addr   = _cfg("NOTIFY_EMAIL_TO")
    smtp_host = _cfg("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(_cfg("SMTP_PORT", "587"))
    smtp_user = _cfg("SMTP_USER")
    smtp_pass = _cfg("SMTP_PASSWORD")
    from_name = _cfg("NOTIFY_FROM_NAME", "Kinevox Academy")

    if not all([to_addr, smtp_user, smtp_pass]):
        print("[email_notify] SMTP not configured — skipping batch doc email.")
        return

    label     = "Employee" if record_type == "employee" else "Student"
    rec_name  = record.get("name", "Unknown")
    rec_id    = record.get("id", "—")
    dept      = record.get("dept") or record.get("course") or "—"
    timestamp = datetime.now().strftime("%d %b %Y at %I:%M %p")
    doc_count = len(documents)

    subject = (
        f"[Kinevox Academy] {doc_count} Document{'s' if doc_count != 1 else ''} Uploaded"
        f" — {rec_name} ({rec_id})"
    )

    # Build rows for each document
    doc_rows_html = ""
    doc_lines_plain = []
    for i, (dtype_label, filepath, orig_name) in enumerate(documents, 1):
        bg = "#f8f8fc" if i % 2 == 1 else "#fff"
        doc_rows_html += f"""
        <tr>
          <td style="padding:9px 14px;font-weight:600;color:#444;background:{bg};
                     border-bottom:1px solid #eee;white-space:nowrap">
            📎 {dtype_label}
          </td>
          <td style="padding:9px 14px;color:#222;border-bottom:1px solid #eee">
            {orig_name}
          </td>
        </tr>"""
        doc_lines_plain.append(f"  [{i}] {dtype_label}: {orig_name}")

    plain_body = (
        f"{doc_count} document(s) uploaded for {label}: {rec_name} ({rec_id})\n"
        f"Uploaded at: {timestamp}\n\n"
        + "\n".join(doc_lines_plain)
        + "\n\nAll files are attached to this email.\n"
        "— Kinevox Academy Management System"
    )

    html_body = f"""<!DOCTYPE html>
<html>
<body style="margin:0;padding:0;background:#f0f0f7;font-family:'DM Sans',Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0">
<tr><td align="center" style="padding:32px 16px">
<table width="560" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:12px;overflow:hidden;
         box-shadow:0 4px 24px rgba(0,0,0,.08)">
  <tr>
    <td style="background:linear-gradient(135deg,#6c63ff,#38bdf8);padding:28px 32px;text-align:center">
      <div style="font-size:22px;font-weight:800;color:#fff">🎓 Kinevox Academy</div>
      <div style="color:rgba(255,255,255,.85);font-size:13px;margin-top:4px">Document Upload Notification</div>
    </td>
  </tr>
  <tr>
    <td style="background:#f0edff;padding:14px 32px;border-bottom:2px solid #6c63ff">
      <span style="font-size:15px;font-weight:700;color:#6c63ff">
        📎 {doc_count} Document{'s' if doc_count != 1 else ''} Uploaded
      </span>
      <span style="font-size:12px;color:#888;float:right;margin-top:2px">{timestamp}</span>
    </td>
  </tr>
  <tr>
    <td style="padding:24px 32px">
      <!-- Person info -->
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border:1px solid #e8e8f0;border-radius:8px;overflow:hidden;margin-bottom:20px">
        <tr>
          <td style="padding:9px 14px;font-weight:600;color:#444;background:#f8f8fc;
                     border-bottom:1px solid #eee;white-space:nowrap">{label} Name</td>
          <td style="padding:9px 14px;color:#222;border-bottom:1px solid #eee">{rec_name}</td>
        </tr>
        <tr>
          <td style="padding:9px 14px;font-weight:600;color:#444;background:#f8f8fc;
                     border-bottom:1px solid #eee;white-space:nowrap">{label} ID</td>
          <td style="padding:9px 14px;color:#222;border-bottom:1px solid #eee">{rec_id}</td>
        </tr>
        <tr>
          <td style="padding:9px 14px;font-weight:600;color:#444;background:#f8f8fc;
                     white-space:nowrap">{'Department' if record_type == 'employee' else 'Institution'}</td>
          <td style="padding:9px 14px;color:#222">{dept}</td>
        </tr>
      </table>

      <!-- Documents list -->
      <div style="font-size:12px;font-weight:700;color:#6c63ff;text-transform:uppercase;
                  letter-spacing:1px;margin-bottom:8px">Uploaded Documents</div>
      <table width="100%" cellpadding="0" cellspacing="0"
             style="border:1px solid #e8e8f0;border-radius:8px;overflow:hidden">
        {doc_rows_html}
      </table>

      <div style="margin-top:16px;padding:14px 18px;background:#f0fff4;
                  border:1px solid #6ee7b7;border-radius:8px;font-size:13px;color:#065f46">
        📎 All {doc_count} file{'s are' if doc_count != 1 else ' is'} attached to this email.
      </div>
    </td>
  </tr>
  <tr>
    <td style="background:#f8f8fc;padding:16px 32px;text-align:center;
               font-size:11px;color:#aaa;border-top:1px solid #eee">
      This is an automated message from Kinevox Academy Management System.
    </td>
  </tr>
</table>
</td></tr>
</table>
</body>
</html>"""

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = f"{from_name} <{smtp_user}>"
    msg["To"]      = to_addr

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(plain_body, "plain"))
    alt.attach(MIMEText(html_body,  "html"))
    msg.attach(alt)

    # Attach every uploaded file
    for dtype_label, filepath, orig_name in documents:
        try:
            mime_type, _ = mimetypes.guess_type(filepath)
            main_type, sub_type = (mime_type or "application/octet-stream").split("/", 1)
            with open(filepath, "rb") as fh:
                part = MIMEBase(main_type, sub_type)
                part.set_payload(fh.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", "attachment", filename=orig_name)
            msg.attach(part)
        except Exception as exc:
            print(f"[email_notify] Could not attach {orig_name}: {exc}")

    try:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=15) as server:
            server.ehlo()
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.sendmail(smtp_user, [to_addr], msg.as_string())
        print(f"[email_notify] ✅ Batch doc email sent — {doc_count} file(s) for {record_type} {rec_id}")
    except Exception as exc:
        print(f"[email_notify] ❌ Batch doc email failed: {exc}")


def notify_documents_batch(
    record_type: str,
    record: dict,
    documents: list,
) -> None:
    """Fire-and-forget: send one email with all uploaded documents attached."""
    t = threading.Thread(
        target=_send_documents_batch_email,
        args=(record_type, record, documents),
        daemon=True,
        name="email-notify-docs-batch",
    )
    t.start()
